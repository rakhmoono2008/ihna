import os
import io
import logging
import time
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, KeyboardButton,
)
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes, ConversationHandler
)

logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# --- НАСТРОЙКИ БОТА ---
BOT_TOKEN  = os.environ.get("BOT_TOKEN",  "8239199632:AAGF3i4Qsswck4WEhaeLqHNvGBgE1bb2uVw")
ADMIN_ID   = int(os.environ.get("ADMIN_ID",   "0"))
ADMIN_ID_2 = int(os.environ.get("ADMIN_ID_2", "0"))
GROUP_ID   = int(os.environ.get("GROUP_ID",   "0"))

# Название видеофайла, который лежит в той же папке, что и скрипт
WELCOME_VIDEO_PATH = "video.mp4" 

# Часовой пояс Ташкента (UTC+5)
TASHKENT_TZ = timezone(timedelta(hours=5))
# ----------------------

LANGUAGE, MENU, APPEAL_TYPE, CONTACT, REGION, APPEAL = range(6)

blocked_users  = {}
user_languages = {}
appeals_db     = {}
appeal_counter = [0]
pending_reply  = {}

CONTACT_USERNAME = "@IHMA_Yoshlar_Yetakchisi"

APPEAL_TYPES = {
    "atype_1": {"ru": "💻 Цифровизация",                       "uz": "💻 Raqamlashtirish",                "kk": "💻 Sanlashtırıw"},
    "atype_2": {"ru": "💰 Стимулирование",                     "uz": "💰 Rag'batlantirish",               "kk": "💰 Ынtalandırıw"},
    "atype_3": {"ru": "📜 Законодательство",                   "uz": "📜 Qonunchilik",                    "kk": "📜 Nızamshılıq"},
    "atype_4": {"ru": "💡 Социальные инновации (новые услуги)", "uz": "💡 Ijtimoiy innovatsiyalar",        "kk": "💡 Sociallıq innovaciyalar"},
    "atype_5": {"ru": "🚗 Логистика и МТО",                    "uz": "🚗 Logistika va MTA",               "kk": "🚗 Logistika"},
    "atype_6": {"ru": "🎓 Повышение квалификации",             "uz": "🎓 Malaka oshirish",                "kk": "🎓 Malasın arttırıw"},
    "atype_7": {"ru": "✨ Другие предложения",                  "uz": "✨ Boshqa takliflar",               "kk": "✨ Basqa usınıslar"},
}

REGIONS = [
    "Toshkent shahri / г. Ташкент", "Toshkent viloyati / Ташкентская обл.",
    "Samarqand / Самарканд", "Farg'ona / Фергана", "Andijon / Андижан",
    "Namangan / Наманган", "Buxoro / Бухара", "Qashqadaryo / Кашкадарья",
    "Surxondaryo / Сурхандарья", "Xorazm / Хорезм", "Navoiy / Навои",
    "Jizzax / Джизак", "Sirdaryo / Сырдарья", "Qoraqalpogiston / Каракалпакстан",
]

TEXTS = {
    "ru": {
        "welcome":        "Добро пожаловать!\n\nЭтот бот предназначен для приёма предложений граждан Узбекистана.\n\nXush kelibsiz!\n\nUshbu bot O'zbekiston fuqarolarining takliflarini qabul qilish uchun mo'ljallangan.\n\nXosh keldińiz!\n\nBul bot Ózbekstan puqaralarınıń usınısların qabıl etiw ushın arnalǵan.\n\nВыберите язык:",
        "menu":           "Главное меню\n\nВыберите действие:",
        "btn_new":        "Новое предложение",
        "btn_my":         "Мои предложения",
        "btn_rules":      "Правила",
        "btn_lang":       "Сменить язык",
        "btn_contacts":   "Контакты",
        "contacts_text":  "📞 Контактные данные\n\nЕсли у вас есть какие-либо вопросы — обратитесь к " + CONTACT_USERNAME,
        "choose_type":    "Выберите тип предложения:",
        "choose_contact": "Хотите оставить свой номер телефона?\n\nВнимание: Если останетесь анонимным — мы не сможем связаться с вами.",
        "share_phone":    "📱 Поделиться номером",
        "stay_anon":      "🙈 Остаться анонимным",
        "anon_warning":   "Вы выбрали анонимность.\nПредложение будет передано, но мы не сможем связаться с вами.",
        "choose_region":  "Выберите ваш регион:",
        "send_appeal":    "Опишите ваше предложение\n\nНапишите текст (можно несколько сообщений подряд).\nФото/видео — отправьте по отдельности.\n\nКогда закончите — нажмите кнопку ниже.",
        "submit_btn":     "Отправить предложение",
        "cancel_btn":     "Отмена",
        "thank_you":      "Ваше предложение передано!\n\nСпасибо за доверие. Мы обязательно рассмотрим ваше предложение.\n\nВы можете следить за ответами в разделе Мои предложения.",
        "media_received": "Файл получен. Добавьте ещё или нажмите Отправить предложение.",
        "text_added":     "Текст добавлен. Продолжайте или нажмите Отправить предложение.",
        "admin_header":   "НОВОЕ ПРЕДЛОЖЕНИЕ",
        "type_label":     "Тип",
        "region_label":   "Регион",
        "contact_label":  "Контакт",
        "anon_label":     "Анонимный",
        "message_label":  "Предложение",
        "no_message":     "(только медиафайлы)",
        "no_appeals":     "У вас пока нет предложений.",
        "your_appeals":   "Ваши предложения:",
        "no_replies":     "Ответов пока нет.",
        "reply_label":    "Ответ",
        "reply_received": "Ответ на ваше предложение #{id}:\n\n{reply}",
        "rules":          "ПРАВИЛА ИСПОЛЬЗОВАНИЯ БОТА\n\n1. Запрещено использовать нецензурные и оскорбительные слова.\n2. Запрещено отправлять спам, рекламу, угрозы.\n3. Запрещено отправлять ложные сведения.\n4. Запрещено злоупотреблять ботом.\n\nКАК ПОЛЬЗОВАТЬСЯ:\n- Нажмите Новое предложение\n- Выберите тип предложения\n- Выберите контакт или анонимность\n- Выберите регион\n- Опишите ситуацию, прикрепите фото/видео\n- Нажмите Отправить предложение\n\nНарушение правил влечёт блокировку.",
        "blocked_perm":   "Вы заблокированы навсегда за нарушение правил.",
        "blocked_temp":   "Вы заблокированы до {until} за нарушение правил.",
        "appeal_id":      "ID предложения",
        "phone_prompt":   "Нажмите кнопку ниже чтобы поделиться номером:",
        "reply_prompt":   "Напишите ответ на предложение #{aid}:",
        "reply_sent":     "Ответ отправлен (#{aid}).",
        "reply_fail":     "Не удалось отправить ответ.",
    },
    "uz": {
        "welcome":        "Xush kelibsiz!\n\nBu bot O'zbekiston fuqarolarining takliflarini qabul qilish uchun.\n\nTilni tanlang:",
        "menu":           "Asosiy menyu\n\nAmalni tanlang:",
        "btn_new":        "Yangi taklif",
        "btn_my":         "Mening takliflarim",
        "btn_rules":      "Qoidalar",
        "btn_lang":       "Tilni o'zgartirish",
        "btn_contacts":   "Kontaktlar",
        "contacts_text":  "📞 Kontakt ma'lumotlari\n\nSavollaringiz bo'lsa — " + CONTACT_USERNAME + " ga murojaat qiling",
        "choose_type":    "Taklif turini tanlang:",
        "choose_contact": "Telefon raqamingizni qoldirmoqchimisiz?\n\nDiqqat: Anonim qolsangiz — biz siz bilan bog'lana olmaymiz.",
        "share_phone":    "📱 Raqamni ulashish",
        "stay_anon":      "🙈 Anonim qolish",
        "anon_warning":   "Siz anonimlikni tanladingiz.\nTaklif uzatiladi, lekin biz siz bilan bog'lana olmaymiz.",
        "choose_region":  "Viloyatingizni tanlang:",
        "send_appeal":    "Taklifingizni yozing\n\nMatn yozing (bir nechta xabar yuborishingiz mumkin).\nRasm/video — alohida yuboring.\n\nTayyor bo'lgach — quyidagi tugmani bosing.",
        "submit_btn":     "Taklifni yuborish",
        "cancel_btn":     "Bekor qilish",
        "thank_you":      "Taklifingiz qabul qilindi!\n\nKatta rahmat. Taklifingizni ko'rib chiqamiz.\n\nMening takliflarim bo'limida javoblarni kuzatishingiz mumkin.",
        "media_received": "Fayl qabul qilindi. Yana qo'shing yoki Taklifni yuborish ni bosing.",
        "text_added":     "Matn qo'shildi. Davom eting yoki Taklifni yuborish ni bosing.",
        "admin_header":   "YANGI TAKLIF",
        "type_label":     "Tur",
        "region_label":   "Viloyat",
        "contact_label":  "Kontakt",
        "anon_label":     "Anonim",
        "message_label":  "Taklif",
        "no_message":     "(faqat media fayllar)",
        "no_appeals":     "Sizda hali takliflar yo'q.",
        "your_appeals":   "Sizning takliflaringiz:",
        "no_replies":     "Hali javob yo'q.",
        "reply_label":    "Javob",
        "reply_received": "#{id}-taklif javobingiz:\n\n{reply}",
        "rules":          "BOTDAN FOYDALANISH QOIDALARI\n\n1. Taqiqlangan: so'kinish, haqoratli so'zlar.\n2. Taqiqlangan: spam, reklama, tahdidlar.\n3. Taqiqlangan: yolg'on ma'lumot.\n4. Taqiqlangan: botni suiiste'mol qilish.\n\nQANDAY FOYDALANISH:\n- Yangi taklif tugmasini bosing\n- Taklif turini tanlang\n- Raqam qoldiring yoki anonim qoling\n- Viloyatni tanlang\n- Muammoni tasvirlang\n- Taklifni yuborish tugmasini bosing\n\nQoidalarni buzish bloklashga olib keladi.",
        "blocked_perm":   "Siz qoidalarni buzganingiz uchun doimiy ravishda bloklandingiz.",
        "blocked_temp":   "Siz {until} gacha bloklangansiz.",
        "appeal_id":      "Taklif ID",
        "phone_prompt":   "Raqamingizni ulashish uchun quyidagi tugmani bosing:",
        "reply_prompt":   "#{aid}-taklif uchun javob yozing:",
        "reply_sent":     "Javob yuborildi (#{aid}).",
        "reply_fail":     "Javob yuborib bo'lmadi.",
    },
    "kk": {
        "welcome":        "Xosh keldiniz!\n\nBul bot O'zbekistan puqaralarinin usınısların qabıl etiw ushin.\n\nTildi tanlang:",
        "menu":           "Bas menyu\n\nAmaldı tanlang:",
        "btn_new":        "Jana usınıs",
        "btn_my":         "Menin usınıslarım",
        "btn_rules":      "Qaydalar",
        "btn_lang":       "Tildi ozgertiw",
        "btn_contacts":   "Kontaktlar",
        "contacts_text":  "📞 Baylanıs maglıwmatları\n\nSawallarınız bolsa — " + CONTACT_USERNAME + " ge murajaat eting",
        "choose_type":    "Usınıs turın tanlang:",
        "choose_contact": "Telefon nomernizdi qaldırganız keleme?\n\nDıqqat: Anonim qalsanız — biz siz benen baylanısa almaymız.",
        "share_phone":    "📱 Nomerni ulasıw",
        "stay_anon":      "🙈 Anonim qalıw",
        "anon_warning":   "Siz anonimlikti tanladınız.\nUsınıs jetkiziledi, biraq biz siz benen baylanısa almaymız.",
        "choose_region":  "Regionınızdı tanlang:",
        "send_appeal":    "Usınısınızdı jazın\n\nMatn jazın (birneshe xabar jibere alasız).\nSuwret/video — boleksha jiberin.\n\nTayar bolganda — to'mendegi tugmasin basın.",
        "submit_btn":     "Usınıs jiberiw",
        "cancel_btn":     "Biykar etiw",
        "thank_you":      "Usınısınız qabıl etildi!\n\nKo'p rahmet. Usınısınızdı ko'rib shıgamız.\n\nMenin usınıslarım boliminde jawaplardı kore alasız.",
        "media_received": "Fayl qabıl etildi. Yana qosın yaki Usınıs jiberiw ni basın.",
        "text_added":     "Matn qosıldı. Davom etin yaki Usınıs jiberiw ni basın.",
        "admin_header":   "JANA USINIS",
        "type_label":     "Tur",
        "region_label":   "Region",
        "contact_label":  "Baylanıs",
        "anon_label":     "Anonim",
        "message_label":  "Usınıs",
        "no_message":     "(tek media fayllar)",
        "no_appeals":     "Sizde halirshe usınıslar joq.",
        "your_appeals":   "Sizin usınıslarınız:",
        "no_replies":     "Halirshe jawap joq.",
        "reply_label":    "Jawap",
        "reply_received": "#{id}-usınıs jawabınız:\n\n{reply}",
        "rules":          "BOTTAN PAYDALANIW QAYDALARI\n\n1. Tıyım: so'kinish, haqoretli so'zler.\n2. Tıyım: spam, reklama, qorqıtıwlar.\n3. Tıyım: jalgan maglıwmat.\n4. Tıyım: bottı qıyanetlep paydalanıw.\n\nQALAYINSHA PAYDALANIW:\n- Jana usınıs tugmasin basın\n- Usınıs turın tanlang\n- Nomer qaldırın yaki anonim qalın\n- Regionınızdı tanlang\n- Maseleni tasvirlen\n- Usınıs jiberiw tugmasin basın\n\nQaydaların bozıw bloklashga alıp keledi.",
        "blocked_perm":   "Siz qaydaların bozganınız ushin mangi bloclandınız.",
        "blocked_temp":   "Siz {until} ga shekem bloclandınız.",
        "appeal_id":      "Usınıs ID",
        "phone_prompt":   "Nomernizdi ulasıw ushin to'mendegi tugmasin basın:",
        "reply_prompt":   "#{aid}-usınıs ushin jawap jazın:",
        "reply_sent":     "Jawap jiberildi (#{aid}).",
        "reply_fail":     "Jawap jiberib bolmadı.",
    },
}


def T(lang, key):
    return TEXTS.get(lang, TEXTS["ru"]).get(key, TEXTS["ru"].get(key, ""))

def get_lang(uid):
    return user_languages.get(uid, "ru")

def is_admin(uid):
    return uid == ADMIN_ID or (ADMIN_ID_2 != 0 and uid == ADMIN_ID_2)

def check_block(uid):
    if uid not in blocked_users:
        return False, ""
    until = blocked_users[uid]
    if until == 0:
        return True, "permanent"
    if time.time() < until:
        # ИСПОЛЬЗУЕМ ВРЕМЯ ТАШКЕНТА ДЛЯ БЛОКИРОВОК
        return True, datetime.fromtimestamp(until, TASHKENT_TZ).strftime("%d.%m.%Y %H:%M")
    del blocked_users[uid]
    return False, ""

def new_aid():
    appeal_counter[0] += 1
    return appeal_counter[0]


def lang_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🇷🇺 Русский", callback_data="lang_ru"),
         InlineKeyboardButton("🇺🇿 O'zbek",  callback_data="lang_uz")],
        [InlineKeyboardButton("🏳️ Qaraqalpaqsha", callback_data="lang_kk")],
    ])

def menu_kb(lang):
    return ReplyKeyboardMarkup([
        [KeyboardButton("✍️ " + T(lang, "btn_new")),   KeyboardButton("📋 " + T(lang, "btn_my"))],
        [KeyboardButton("📜 " + T(lang, "btn_rules")), KeyboardButton("📞 " + T(lang, "btn_contacts"))],
        [KeyboardButton("🌐 " + T(lang, "btn_lang"))],
    ], resize_keyboard=True)

def appeal_type_kb(lang):
    rows = []
    for key, names in APPEAL_TYPES.items():
        rows.append([InlineKeyboardButton(names.get(lang, names["ru"]), callback_data=key)])
    return InlineKeyboardMarkup(rows)

def contact_kb(lang):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📱 " + T(lang, "share_phone"), callback_data="contact_phone")],
        [InlineKeyboardButton("🙈 " + T(lang, "stay_anon"),   callback_data="contact_anon")],
    ])

def region_kb():
    rows = []
    for i in range(0, len(REGIONS), 2):
        row = [InlineKeyboardButton(REGIONS[i], callback_data="reg_" + str(i))]
        if i + 1 < len(REGIONS):
            row.append(InlineKeyboardButton(REGIONS[i+1], callback_data="reg_" + str(i+1)))
        rows.append(row)
    return InlineKeyboardMarkup(rows)

def submit_kb(lang):
    return ReplyKeyboardMarkup([
        [KeyboardButton("✅ " + T(lang, "submit_btn"))],
        [KeyboardButton("❌ " + T(lang, "cancel_btn"))],
    ], resize_keyboard=True)

def admin_kb(aid, uid):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 Ответить на предложение", callback_data="reply_" + str(aid) + "_" + str(uid))],
        [InlineKeyboardButton("⏱ 1 час",    callback_data="block_" + str(uid) + "_1h"),
         InlineKeyboardButton("⏱ 3 часа",   callback_data="block_" + str(uid) + "_3h")],
        [InlineKeyboardButton("📅 1 день",   callback_data="block_" + str(uid) + "_1d"),
         InlineKeyboardButton("📅 1 неделя", callback_data="block_" + str(uid) + "_1w")],
        [InlineKeyboardButton("🚫 Заблокировать навсегда", callback_data="block_" + str(uid) + "_perm")],
    ])


async def send_to_targets(bot, caption, media_ids, kb=None):
    targets = [ADMIN_ID]
    if ADMIN_ID_2:
        targets.append(ADMIN_ID_2)
    if GROUP_ID:
        targets.append(GROUP_ID)
    targets = list(dict.fromkeys(targets))
    send_map = {"photo": bot.send_photo, "video": bot.send_video, "document": bot.send_document}
    for target in targets:
        try:
            if media_ids:
                fn = send_map.get(media_ids[0]["type"])
                if fn:
                    await fn(target, media_ids[0]["id"], caption=caption, reply_markup=kb)
                for m in media_ids[1:]:
                    fn2 = send_map.get(m["type"])
                    if fn2:
                        await fn2(target, m["id"])
            else:
                await bot.send_message(target, caption, reply_markup=kb)
        except Exception as e:
            logger.error("Send to " + str(target) + " failed: " + str(e))


async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    blocked, until = check_block(uid)
    if blocked:
        lang = get_lang(uid)
        msg = T(lang, "blocked_perm") if until == "permanent" else T(lang, "blocked_temp").format(until=until)
        await update.message.reply_text(msg)
        return ConversationHandler.END
    ctx.user_data.clear()

    # ОТПРАВКА ВИДЕО ПРИ СТАРТЕ
    try:
        if os.path.exists(WELCOME_VIDEO_PATH):
            with open(WELCOME_VIDEO_PATH, "rb") as video_file:
                await update.message.reply_video(
                    video=video_file,
                    caption=TEXTS["ru"]["welcome"],
                    reply_markup=lang_kb()
                )
        else:
            # Если видео нет, просто отправляем текст
            await update.message.reply_text(TEXTS["ru"]["welcome"], reply_markup=lang_kb())
    except Exception as e:
        logger.error(f"Ошибка при отправке видео: {e}")
        await update.message.reply_text(TEXTS["ru"]["welcome"], reply_markup=lang_kb())

    return LANGUAGE


async def on_language(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    lang = {"lang_ru": "ru", "lang_uz": "uz", "lang_kk": "kk"}.get(q.data, "ru")
    ctx.user_data["lang"] = lang
    user_languages[q.from_user.id] = lang
    await q.edit_message_reply_markup(reply_markup=None)
    await q.message.reply_text(T(lang, "menu"), reply_markup=menu_kb(lang))
    return MENU


async def on_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    lang = ctx.user_data.get("lang", get_lang(uid))
    text = update.message.text

    blocked, until = check_block(uid)
    if blocked:
        msg = T(lang, "blocked_perm") if until == "permanent" else T(lang, "blocked_temp").format(until=until)
        await update.message.reply_text(msg)
        return MENU

    if T(lang, "btn_new") in text:
        await update.message.reply_text(T(lang, "choose_type"), reply_markup=appeal_type_kb(lang))
        return APPEAL_TYPE

    if T(lang, "btn_my") in text:
        user_appeals = [(k, v) for k, v in appeals_db.items() if v["user_id"] == uid]
        if not user_appeals:
            await update.message.reply_text(T(lang, "no_appeals"))
            return MENU
        await update.message.reply_text(T(lang, "your_appeals"))
        for aid, ap in sorted(user_appeals):
            preview = (ap.get("text") or "")[:80]
            appeal_type = ap.get("appeal_type", "")
            lines = [
                "#" + str(aid) + " | " + ap["date"],
                T(lang, "type_label") + ": " + appeal_type,
                ap["region"],
                preview or T(lang, "no_message"),
            ]
            if ap.get("replies"):
                for r in ap["replies"]:
                    lines.append("")
                    lines.append(T(lang, "reply_label") + ": " + r)
            else:
                lines.append("")
                lines.append(T(lang, "no_replies"))
            await update.message.reply_text("\n".join(lines))
        return MENU

    if T(lang, "btn_rules") in text:
        await update.message.reply_text(T(lang, "rules"))
        return MENU

    if T(lang, "btn_contacts") in text:
        await update.message.reply_text(T(lang, "contacts_text"))
        return MENU

    if T(lang, "btn_lang") in text:
        await update.message.reply_text(TEXTS["ru"]["welcome"], reply_markup=lang_kb())
        return LANGUAGE

    return MENU


async def on_appeal_type(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    lang = ctx.user_data.get("lang", get_lang(q.from_user.id))
    appeal_type = APPEAL_TYPES.get(q.data, {}).get(lang, q.data)
    ctx.user_data["appeal_type"] = appeal_type
    await q.edit_message_text("✅ " + appeal_type)
    await q.message.reply_text(T(lang, "choose_contact"), reply_markup=contact_kb(lang))
    return CONTACT


async def on_contact(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    lang = ctx.user_data.get("lang", get_lang(q.from_user.id))
    if q.data == "contact_phone":
        phone_kb = ReplyKeyboardMarkup(
            [[KeyboardButton("📱 " + T(lang, "share_phone"), request_contact=True)]],
            resize_keyboard=True, one_time_keyboard=True
        )
        await q.edit_message_text(T(lang, "phone_prompt"))
        await q.message.reply_text("👇", reply_markup=phone_kb)
        return REGION
    else:
        ctx.user_data["contact"] = None
        await q.edit_message_text(T(lang, "anon_warning"))
        await q.message.reply_text(T(lang, "choose_region"), reply_markup=region_kb())
        return REGION


async def on_phone(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    lang = ctx.user_data.get("lang", get_lang(update.effective_user.id))
    ctx.user_data["contact"] = update.message.contact.phone_number
    await update.message.reply_text(T(lang, "choose_region"), reply_markup=region_kb())
    return REGION


async def on_region(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    lang = ctx.user_data.get("lang", get_lang(q.from_user.id))
    idx  = int(q.data.split("_")[1])
    ctx.user_data["region"]     = REGIONS[idx]
    ctx.user_data["text_parts"] = []
    ctx.user_data["media_ids"]  = []
    await q.edit_message_text("✅ " + REGIONS[idx])
    await q.message.reply_text(T(lang, "send_appeal"), reply_markup=submit_kb(lang))
    return APPEAL


async def on_appeal_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    lang = ctx.user_data.get("lang", get_lang(update.effective_user.id))
    text = update.message.text
    if "✅ " + T(lang, "submit_btn") in text:
        return await do_submit(update, ctx)
    if "❌ " + T(lang, "cancel_btn") in text:
        ctx.user_data["lang"] = lang
        await update.message.reply_text(T(lang, "menu"), reply_markup=menu_kb(lang))
        return MENU
    ctx.user_data.setdefault("text_parts", []).append(text)
    await update.message.reply_text(T(lang, "text_added"))
    return APPEAL


async def on_appeal_media(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    lang      = ctx.user_data.get("lang", get_lang(update.effective_user.id))
    media_ids = ctx.user_data.setdefault("media_ids", [])
    if update.message.photo:
        media_ids.append({"type": "photo", "id": update.message.photo[-1].file_id})
    elif update.message.video:
        media_ids.append({"type": "video", "id": update.message.video.file_id})
    elif update.message.document:
        media_ids.append({"type": "document", "id": update.message.document.file_id})
    if update.message.caption:
        ctx.user_data.setdefault("text_parts", []).append(update.message.caption)
    await update.message.reply_text(T(lang, "media_received"))
    return APPEAL


async def do_submit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    lang        = ctx.user_data.get("lang", get_lang(update.effective_user.id))
    region      = ctx.user_data.get("region", "—")
    contact     = ctx.user_data.get("contact")
    appeal_type = ctx.user_data.get("appeal_type", "—")
    text_parts  = ctx.user_data.get("text_parts", [])
    full_text   = "\n\n".join(text_parts) if text_parts else ""
    media_ids   = ctx.user_data.get("media_ids", [])
    user        = update.effective_user
    aid         = new_aid()
    
    # ИСПОЛЬЗУЕМ ВРЕМЯ ТАШКЕНТА ПРИ СОХРАНЕНИИ ДАТЫ
    date_str    = datetime.now(TASHKENT_TZ).strftime("%d.%m.%Y %H:%M")

    appeals_db[aid] = {
        "user_id": user.id, "user_name": user.full_name, "region": region, "contact": contact,
        "text": full_text, "media": media_ids, "date": date_str,
        "replies": [], "lang": lang, "appeal_type": appeal_type,
    }

    contact_str = ("+" + contact) if contact else T(lang, "anon_label")
    lang_names  = {"ru": "Русский", "uz": "O'zbek", "kk": "Qaraqalpaqsha"}
    caption = "\n".join([
        T(lang, "admin_header"),
        "",
        "Пользователь: " + user.full_name + " (@" + (user.username or "—") + ") [" + str(user.id) + "]",
        T(lang, "appeal_id") + ": #" + str(aid),
        T(lang, "type_label") + ": " + appeal_type,
        T(lang, "region_label") + ": " + region,
        T(lang, "contact_label") + ": " + contact_str,
        "Язык: " + lang_names.get(lang, lang),
        "Дата: " + date_str,
        "",
        T(lang, "message_label") + ":",
        full_text if full_text else T(lang, "no_message"),
    ])
    await send_to_targets(update.get_bot(), caption, media_ids, admin_kb(aid, user.id))
    await update.message.reply_text(T(lang, "thank_you"), reply_markup=menu_kb(lang))
    ctx.user_data.clear()
    ctx.user_data["lang"] = lang
    return MENU


async def on_admin_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if not is_admin(q.from_user.id):
        await q.answer("Нет доступа", show_alert=True)
        return
    await q.answer()
    data      = q.data
    admin_uid = q.from_user.id

    if data.startswith("reply_"):
        parts = data.split("_")
        aid   = int(parts[1])
        uid   = int(parts[2])
        pending_reply[admin_uid] = {"appeal_id": aid, "user_id": uid}
        try:
            await q.edit_message_reply_markup(reply_markup=None)
        except Exception:
            pass
        await q.message.reply_text(T("ru", "reply_prompt").format(aid=aid))
        return

    if data.startswith("block_"):
        parts     = data.split("_")
        uid       = int(parts[1])
        dur       = parts[2]
        durations = {"1h": 3600, "3h": 10800, "1d": 86400, "1w": 604800, "perm": 0}
        labels    = {"1h": "1 час", "3h": "3 часа", "1d": "1 день", "1w": "1 неделю", "perm": "навсегда"}
        secs      = durations.get(dur, 3600)
        label     = labels.get(dur, dur)
        if secs == 0:
            blocked_users[uid] = 0
            until_str = "навсегда"
        else:
            until_ts = time.time() + secs
            blocked_users[uid] = until_ts
            # ИСПОЛЬЗУЕМ ВРЕМЯ ТАШКЕНТА ДЛЯ ОТОБРАЖЕНИЯ СРОКА
            until_str = datetime.fromtimestamp(until_ts, TASHKENT_TZ).strftime("%d.%m.%Y %H:%M")
        user_lang = user_languages.get(uid, "ru")
        try:
            msg = T(user_lang, "blocked_perm") if secs == 0 else T(user_lang, "blocked_temp").format(until=until_str)
            await ctx.bot.send_message(uid, msg)
        except Exception as e:
            logger.error("Block notify failed: " + str(e))
        try:
            await q.edit_message_reply_markup(reply_markup=None)
        except Exception:
            pass
        await q.message.reply_text("Пользователь " + str(uid) + " заблокирован на " + label + ".")


async def on_admin_message(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not is_admin(uid):
        return
    if uid not in pending_reply:
        return
    text       = update.message.text or ""
    info       = pending_reply.pop(uid)
    aid        = info["appeal_id"]
    target_uid = info["user_id"]
    user_lang  = appeals_db.get(aid, {}).get("lang", user_languages.get(target_uid, "ru"))
    if aid in appeals_db:
        appeals_db[aid]["replies"].append(text)
    try:
        await ctx.bot.send_message(
            target_uid,
            T(user_lang, "reply_received").format(id=aid, reply=text)
        )
        await update.message.reply_text(T("ru", "reply_sent").format(aid=aid))
    except Exception as e:
        logger.error("Reply send failed: " + str(e))
        await update.message.reply_text(T("ru", "reply_fail"))


def build_excel(appeal_list, title="Отчёт"):
    """Build Excel file from list of appeals and return as BytesIO."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Предложения"

    # Styles
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="AAAAAA")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "№", "ID", "Дата", "Тип предложения", "Регион",
        "Контакт", "Язык", "Пользователь", "Telegram ID",
        "Текст предложения", "Ответы"
    ]
    col_widths = [5, 7, 18, 28, 30, 18, 12, 22, 14, 50, 40]

    # Header row
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font        = header_font
        cell.fill        = header_fill
        cell.alignment   = center_align
        cell.border      = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22

    # Type colors
    type_colors = {
        "Цифровизация":          "DDEEFF",
        "Стимулирование":        "DDFFDD",
        "Законодательство":      "FFEECC",
        "Социальные инновации":  "FFDDE8",
        "Логистика":             "E8DDFF",
        "Повышение квалификации":"FFFACC",
        "Другие":                "F0F0F0",
    }

    lang_names = {"ru": "Русский", "uz": "O'zbek", "kk": "Qaraqalpaqsha"}

    for row_num, (aid, ap) in enumerate(sorted(appeal_list), 1):
        row = row_num + 1
        contact = ("+" + ap["contact"]) if ap.get("contact") else "Анонимный"
        replies = "; ".join(ap.get("replies", [])) if ap.get("replies") else "—"
        atype   = ap.get("appeal_type", "—")

        # Pick row fill based on type keyword
        fill_color = "FFFFFF"
        for keyword, color in type_colors.items():
            if keyword.lower() in atype.lower():
                fill_color = color
                break
        row_fill = PatternFill("solid", fgColor=fill_color)

        values = [
            row_num,
            "#" + str(aid),
            ap.get("date", ""),
            atype,
            ap.get("region", ""),
            contact,
            lang_names.get(ap.get("lang", "ru"), "Русский"),
            ap.get("user_name", "—"),
            str(ap.get("user_id", "")),
            ap.get("text", "") or "—",
            replies,
        ]
        aligns = [center_align, center_align, center_align, left_align, left_align,
                  center_align, center_align, left_align, center_align, left_align, left_align]

        for col_idx, (val, aln) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = aln
            cell.border    = border
            cell.fill      = row_fill

        ws.row_dimensions[row].height = 40

    # Summary sheet
    ws2 = wb.create_sheet("Итоги")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 12

    h1 = ws2.cell(row=1, column=1, value="Категория")
    h2 = ws2.cell(row=1, column=2, value="Количество")
    for c in [h1, h2]:
        c.font      = header_font
        c.fill      = header_fill
        c.alignment = center_align
        c.border    = border

    type_counts = {}
    for aid, ap in appeal_list:
        t = ap.get("appeal_type", "—")
        type_counts[t] = type_counts.get(t, 0) + 1

    for r, (t, cnt) in enumerate(sorted(type_counts.items()), 2):
        ws2.cell(row=r, column=1, value=t).border   = border
        ws2.cell(row=r, column=2, value=cnt).border = border
        ws2.cell(row=r, column=2).alignment = center_align

    total_row = len(type_counts) + 2
    tc = ws2.cell(row=total_row, column=1, value="ИТОГО")
    vc = ws2.cell(row=total_row, column=2, value=len(appeal_list))
    for c in [tc, vc]:
        c.font   = Font(bold=True)
        c.border = border
    vc.alignment = center_align

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def on_report_command(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not is_admin(uid):
        return
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Отчёт за сегодня", callback_data="report_today")],
        [InlineKeyboardButton("📊 Полный отчёт",     callback_data="report_all")],
    ])
    await update.message.reply_text("Выберите тип отчёта:", reply_markup=kb)


async def on_report_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if not is_admin(q.from_user.id):
        await q.answer("Нет доступа", show_alert=True)
        return
    await q.answer()
    await q.edit_message_reply_markup(reply_markup=None)

    # ИСПОЛЬЗУЕМ ВРЕМЯ ТАШКЕНТА ДЛЯ ОТЧЕТА ПО ТЕКУЩЕМУ ДНЮ
    today_str = datetime.now(TASHKENT_TZ).strftime("%d.%m.%Y")
    
    if q.data == "report_today":
        filtered = [(aid, ap) for aid, ap in appeals_db.items()
                    if ap.get("date", "").startswith(today_str)]
        filename = "report_today_" + today_str.replace(".", "-") + ".xlsx"
        caption  = "📅 Отчёт за сегодня (" + today_str + "): " + str(len(filtered)) + " предложений"
    else:
        filtered = list(appeals_db.items())
        filename = "report_full_" + today_str.replace(".", "-") + ".xlsx"
        caption  = "📊 Полный отчёт: " + str(len(filtered)) + " предложений"

    if not filtered:
        await q.message.reply_text("Предложений не найдено.")
        return

    await q.message.reply_text("⏳ Формирую Excel файл...")
    buf = build_excel(filtered)
    await ctx.bot.send_document(
        chat_id=q.from_user.id,
        document=buf,
        filename=filename,
        caption=caption,
    )

async def on_error(update: object, ctx: ContextTypes.DEFAULT_TYPE):
    logger.error("Error:", exc_info=ctx.error)


def main():
    app = Application.builder().token(BOT_TOKEN).build()
    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            LANGUAGE:    [CallbackQueryHandler(on_language, pattern="^lang_")],
            MENU:        [MessageHandler(filters.TEXT & ~filters.COMMAND, on_menu)],
            APPEAL_TYPE: [CallbackQueryHandler(on_appeal_type, pattern="^atype_")],
            CONTACT:     [CallbackQueryHandler(on_contact, pattern="^contact_")],
            REGION:      [
                CallbackQueryHandler(on_region, pattern="^reg_"),
                MessageHandler(filters.CONTACT, on_phone),
            ],
            APPEAL:      [
                MessageHandler(filters.TEXT & ~filters.COMMAND, on_appeal_text),
                MessageHandler(filters.PHOTO | filters.VIDEO | filters.Document.ALL, on_appeal_media),
            ],
        },
        fallbacks=[CommandHandler("cancel", lambda u, c: ConversationHandler.END)],
        allow_reentry=True,
    )
    app.add_handler(conv)
    app.add_handler(CallbackQueryHandler(on_admin_callback, pattern="^(reply_|block_)"))
    admin_filter = filters.User(ADMIN_ID)
    if ADMIN_ID_2:
        admin_filter = admin_filter | filters.User(ADMIN_ID_2)
    app.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND & admin_filter,
        on_admin_message
    ), group=1)
    app.add_handler(CommandHandler("report", on_report_command))
    app.add_handler(CallbackQueryHandler(on_report_callback, pattern="^report_"))
    app.add_error_handler(on_error)
    logger.info("Bot running...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
